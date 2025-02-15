from app import application, db
from app.Models.Agents import *
from app.Models.Logins import *
from app.Models.VoterDetails import *
from app.Authentication.jwtservice import JWTService
from app.Authentication.middleware import Middleware
from app.Authentication.hashingservice import HashingService
from flask import request, Blueprint, redirect, url_for,jsonify
from werkzeug import exceptions
import csv, io
from openpyxl import load_workbook
from io import StringIO
import pandas as pd
sign_up_key = "signupkey"
jwt_secret = "secret"

jwt_service = JWTService(jwt_secret)
middleware = Middleware(jwt_service)
hashing_service = HashingService()

application.before_request(lambda: middleware.auth(request))

Agents_API_blueprint = Blueprint("Agents_API", __name__)


#signup
@Agents_API_blueprint.route("/agent/signup", methods=["POST"])
def sign_up():
    (
        First_name,
        Last_name,
        Username,
        Password,
        Email_Id,
        IsAdmin,
        Gender,
        Phone_No,
        Address,
    ) = (
        request.json["First_name"],
        request.json["Last_name"],
        request.json["Username"],
        request.json["Password"],
        request.json["Email_Id"],
        0,
        request.json["Gender"],
        request.json["Phone_No"],
        request.json["Address"],
    )
     
    print(f'request.headers.get("sign_up_key"): {request.headers.get("signupkey")}')
    print(f"sign_up_key: {sign_up_key}")
    if request.headers.get("signupkey") != sign_up_key:
        return exceptions.Unauthorized(description="Incorrect Key")
    password_hash = hashing_service.hash_bcrypt(Password.encode("utf-8")).decode(
        "utf-8"
    )

    agent = Agents(
        First_name,
        Last_name,
        Username,
        password_hash,
        Email_Id,
        0,
        Gender,
        Phone_No,
        Address,
    )
    db.session.add(agent)
    db.session.commit()
    return {"message": "Agent Id created Successfully"}

# login start
@Agents_API_blueprint.route("/agent/login", methods=["PUT"])
def log_in():
    username, password = request.json["Username"], request.json["Password"]
    agent = Agents.query.filter_by(Username=username).filter_by(IsAdmin=0).first()
    if agent is None:
        return exceptions.Unauthorized(
            description="Invalid credentials"
        )
    is_password_correct = hashing_service.check_bcrypt(
        password.encode("utf-8"), agent.Hash_Password.encode("utf-8")
    )

    if not is_password_correct:
        return exceptions.Unauthorized(
            description="Incorrect username/password"
        )
    token_payload = {"username": username}
    token = jwt_service.generate(token_payload)

    if token is None:
        return exceptions.InternalServerError(description="Login Failed")
        
    login = Logins.query.filter_by(User_Id=agent.Agent_Id).first()

    if not login:
        print("no record")
        user_id = agent.Agent_Id
        ip_address = request.remote_addr
        device = request.user_agent.string
        status = "LoggedIn"

        # Create a new login entry
        login = Logins(
            User_Id=user_id,
            IP_Address=ip_address,
            Device=device,
            Token=token,
            Status=status,
        )
        db.session.add(login)
        db.session.commit()
    else:
        print("There is record")
        print(token)
        login.Token = token
        print(login.Token)
        login.Status = "LoggedIn"
        db.session.commit()

    # Add the login entry to the session and commit the changes
    return {"token": token}
#login end

@Agents_API_blueprint.route("/agent/is_logged_in")
def is_logged_in():
    return {"message": "token is valid"}


@Agents_API_blueprint.route("/agent/change_password", methods=["POST"])
def change_password():
    token = request.headers["token"]
    username, old_password, new_password, retype_new_password = (
        request.json["Username"],
        request.json["Old_Password"],
        request.json["New_Password"],
        request.json["Retype_New_Password"]
    )
    
    if new_password != retype_new_password:
        return exceptions.Unauthorized(description="Password mismatch")
    agent = Agents.query.filter_by(Username=username).first()
    if agent is None:
        redirect(url_for("Agents_API.agent_log_out", token = token))
        return exceptions.Unauthorized(description="Incorrect username")
    is_password_correct = hashing_service.check_bcrypt(
        old_password.encode("utf-8"), agent.Hash_Password.encode("utf-8")
    )

    if not is_password_correct:
        redirect(url_for("Agents_API.agent_log_out", token = token))
        return exceptions.Unauthorized(description="Incorrect password")
    agent.Hash_Password = hashing_service.hash_bcrypt(
        new_password.encode("utf-8")
    ).decode("utf-8")
    db.session.commit()
    print("password changed")
    print(token)
    redirect(url_for("Agents_API.agent_log_out", token = token,_method = "PUT"))
    print("after logout method")
    return {"message": "Password changed successfully"}


@Agents_API_blueprint.route("/agent/logout/", methods=["PUT"])
def agent_log_out():
    try:
        print("triggered log_out")
        token = request.args.get('token')
        # token = request.headers["token"]
        login = Logins.query.filter_by(Token = token).first()
        login.Status = "LoggedOut"
        db.session.commit()
        print("logged out")
    except Exception as e:
        print(str(e))
        db.session.rollback()
        return jsonify("Error: " + str(e))
    return {"message": "Logged out successfully"}


@Agents_API_blueprint.route("/agent/upload_voterdetails", methods=["POST"])
def upload_data():
    file = request.files["file"]
    if file.filename == '':
        return  {"message": "File not selected"}
    if file:
        try:         
            workbook = load_workbook(file, data_only=True)
            sheet = workbook.active
            #sheet = workbook["Sheet1"] # to access sheet1 or replace with name of sheet

            headers = [cell.value for cell in sheet[1]]  # Assuming the headers are in the first row

            records = []
            for row in sheet.iter_rows(values_only=True, min_row=2):
                record = dict(zip(headers, row))
                record.pop('Voter_Details_Id', None)
                records.append(record)

            
            print("No of records",len(records))
            for record in records:
                data = VoterDetails(**record)
                db.session.add(data)
                db.session.commit()

            #db.session.add(bulk_data)
            # db.session.commit()
            return {"message": "File uploaded successfully"}
            
        except Exception as e:
            db.session.rollback()
            return jsonify("Error: " + str(e))            